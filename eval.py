import os
import os.path as osp
import argparse
from mmengine.runner import Runner
from mmengine.config import Config, DictAction

import vlex_sam3_segmentor
import custom_datasets


def parse_args():
    parser = argparse.ArgumentParser(
        description='VLex-SAM3 evaluation with MMSeg')
    parser.add_argument(
        'config',
        nargs='?',
        default='./configs/cfg_vaihingen.py',
        help='config file path',
    )
    parser.add_argument(
        '--show', action='store_true', help='show prediction results')
    parser.add_argument(
        '--show_dir',
        default='./show_dir/',
        help='directory to save visualizaion images')
    parser.add_argument(
        '--out',
        type=str,
        help='The directory to save output prediction for offline evaluation')
    parser.add_argument(
        '--save-xlsx',
        type=str,
        default=None,
        help='Optional Excel file used to append evaluation metrics.')
    parser.add_argument(
        '--wait-time',
        type=float,
        default=2,
        help='Display interval when --show is enabled.')
    parser.add_argument(
        '--cfg-options',
        nargs='+',
        action=DictAction,
        help='override some settings in the used config, the key-value pair '
        'in xxx=yyy format will be merged into config file. If the value to '
        'be overwritten is a list, it should be like key="[a,b]" or key=a,b '
        'It also allows nested list/tuple values, e.g. key="[(a,b),(c,d)]" '
        'Note that the quotation marks are necessary and that no white space '
        'is allowed.')
    parser.add_argument(
        '--launcher',
        choices=['none', 'pytorch', 'slurm', 'mpi'],
        default='none',
        help='job launcher')
    # When using PyTorch version >= 2.0.0, the `torch.distributed.launch`
    # will pass the `--local-rank` parameter to `tools/train.py` instead
    # of `--local_rank`.
    parser.add_argument('--local_rank', '--local-rank', type=int, default=0)
    args = parser.parse_args()
    if 'LOCAL_RANK' not in os.environ:
        os.environ['LOCAL_RANK'] = str(args.local_rank)

    return args


def append_experiment_result(file_path, experiment_data):
    try:
        import openpyxl
    except ImportError:
        print('Warning: openpyxl is not installed. Skip writing Excel metrics.')
        return

    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if sheet['A1'].value is None:
        sheet['A1'] = 'Model'
        sheet['B1'] = 'Dataset'
        sheet['C1'] = 'aAcc'
        sheet['D1'] = 'mIoU'
        sheet['E1'] = 'mAcc'

    last_row = sheet.max_row

    for index, result in enumerate(experiment_data, start=1):
        sheet.cell(row=last_row + index, column=1, value=result['Model'])
        sheet.cell(row=last_row + index, column=2, value=result['Dataset'])
        sheet.cell(row=last_row + index, column=3, value=result['aAcc'])
        sheet.cell(row=last_row + index, column=4, value=result['mIoU'])
        sheet.cell(row=last_row + index, column=5, value=result['mAcc'])

    workbook.save(file_path)


def trigger_visualization_hook(cfg, args):
    default_hooks = cfg.default_hooks
    if 'visualization' in default_hooks:
        visualization_hook = default_hooks['visualization']
        # Turn on visualization
        visualization_hook['draw'] = True
        if args.show:
            visualization_hook['show'] = True
            visualization_hook['wait_time'] = args.wait_time
        if args.show_dir:
            visualizer = cfg.visualizer
            visualizer['save_dir'] = args.show_dir
    else:
        raise RuntimeError(
            'VisualizationHook must be included in default_hooks.'
            'refer to usage '
            '"visualization=dict(type=\'VisualizationHook\')"')

    return cfg


def main():
    args = parse_args()
    cfg = Config.fromfile(args.config)
    cfg.launcher = args.launcher
    # add output_dir in metric
    if args.out is not None:
        cfg.test_evaluator['output_dir'] = args.out
        cfg.test_evaluator['keep_results'] = True
    if args.cfg_options is not None:
        cfg.merge_from_dict(args.cfg_options)
    if not hasattr(cfg, 'work_dir') or cfg.work_dir is None:
        cfg.work_dir = osp.join('./work_dirs',
                                osp.splitext(osp.basename(args.config))[0])

    # trigger_visualization_hook(cfg, args)
    runner = Runner.from_cfg(cfg)
    results = runner.test()

    results.update({'Model': cfg.model.model_type,
                    'Dataset': cfg.dataset_type})

    if args.save_xlsx is not None and runner.rank == 0:
        append_experiment_result(args.save_xlsx, [results])

    if runner.rank == 0:
        with open(os.path.join(cfg.work_dir, 'results.txt'), 'a') as f:
            f.write(os.path.basename(args.config).split('.')[0] + '\n')
            for k, v in results.items():
                f.write(k + ': ' + str(v) + '\n')


if __name__ == '__main__':
    main()
